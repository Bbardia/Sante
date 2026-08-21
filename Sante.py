import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
from datetime import datetime
import os
import sys
import platform
import pandas as pd

# ================= DATABASE =================

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

db_path = os.path.join(BASE_DIR, "database.db")

conn = sqlite3.connect(db_path)

c = conn.cursor()

# ADD total_value COLUMN IF NOT EXISTS
try:
    c.execute("ALTER TABLE inventory ADD COLUMN total_value REAL DEFAULT 0")
    conn.commit()
except sqlite3.OperationalError:
    pass  # column already exists

c.execute("""
CREATE TABLE IF NOT EXISTS inventory(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT UNIQUE,
    qty REAL NOT NULL,
    unit TEXT,
    total_value REAL NOT NULL DEFAULT 0
)
""")

c.execute("""
CREATE TABLE IF NOT EXISTS products(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT UNIQUE,
    price REAL
)
""")

c.execute("""
CREATE TABLE IF NOT EXISTS recipes(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    product TEXT,
    ingredient TEXT,
    qty REAL
)
""")

c.execute("""
CREATE TABLE IF NOT EXISTS sales(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    product TEXT,
    qty REAL,
    total REAL,
    date TEXT,
    customer TEXT,
    payment_status TEXT
)
""")

# --- Ensure new columns exist in sales table ---

try:
    c.execute("ALTER TABLE sales ADD COLUMN customer TEXT")
except:
    pass  # column already exists

try:
    c.execute("ALTER TABLE sales ADD COLUMN payment_status TEXT")
except:
    pass  # column already exists

conn.commit()


c.execute("""
            CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE,
            discount REAL DEFAULT 0
)
""")
        

conn.commit()

# ================= ROOT =================

root = tk.Tk()
root.title("Santé")
root.geometry("1400x850")

notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True)

frames = {}

for tab in ["Inventory", "Products", "Recipes", "Sales", "Debts", "Reports"]:
    frame = ttk.Frame(notebook)
    notebook.add(frame, text=tab)
    frames[tab] = frame

history_tab = ttk.Frame(notebook)
notebook.add(history_tab, text="Sales History")    

# =========================================================
# INVENTORY TAB
# =========================================================

inventory_frame = frames["Inventory"]

inventory_form = ttk.Frame(inventory_frame)
inventory_form.pack(pady=10)

ttk.Label(inventory_form, text="Ingredient").grid(row=0, column=0, padx=5, pady=5)
ttk.Label(inventory_form, text="Quantity").grid(row=1, column=0, padx=5, pady=5)
ttk.Label(inventory_form, text="Unit").grid(row=2, column=0, padx=5, pady=5)
ttk.Label(inventory_form, text="Price").grid(row=3, column=0, padx=5, pady=5)

inv_name_entry = ttk.Combobox(
    inventory_form,
    width=30,
    state="normal"   # allows typing AND selecting
)
inv_name_entry.grid(row=0, column=1, pady=5)
inv_qty_entry = ttk.Entry(inventory_form, width=30)
inv_unit_entry = ttk.Entry(inventory_form, width=30)
inv_price_entry = ttk.Entry(inventory_form, width=30)

inv_name_entry.grid(row=0, column=1, pady=5)
inv_qty_entry.grid(row=1, column=1, pady=5)
inv_unit_entry.grid(row=2, column=1, pady=5)
inv_price_entry.grid(row=3, column=1, pady=5)

inventory_tree = ttk.Treeview(
    inventory_frame,
    columns=("ID", "Name", "Qty", "Unit", "Avg Price"),
    show="headings",
    height=20
)

for col in ("ID", "Name", "Qty", "Unit", "Avg Price"):
    inventory_tree.heading(col, text=col)

inventory_tree.pack(fill="both", expand=True, padx=10, pady=10)

def refresh_inventory():

    for row in inventory_tree.get_children():
        inventory_tree.delete(row)

    c.execute("""
        SELECT id, name, qty, unit,
        CASE 
            WHEN qty > 0 THEN ROUND(total_value / qty, 2)
            ELSE 0
        END AS avg_price
        FROM inventory
        ORDER BY id
    """)

    for row in c.fetchall():
        inventory_tree.insert("", "end", values=row)
        
def refresh_recipe_combos():

    ingredients = [
        r[0]
        for r in c.execute(
            "SELECT name FROM inventory ORDER BY name"
        ).fetchall()
    ]

    products = [
        r[0]
        for r in c.execute(
            "SELECT name FROM products ORDER BY name"
        ).fetchall()
    ]

    recipe_ingredient_combo["values"] = ingredients
    recipe_product_combo["values"] = products

def refresh_sales_combo():

    products = [
        r[0]
        for r in c.execute(
            "SELECT name FROM products ORDER BY name"
        ).fetchall()
    ]

    sales_product_combo["values"] = products

def refresh_inventory_combo():
    names = [
        r[0]
        for r in c.execute(
            "SELECT name FROM inventory ORDER BY name"
        ).fetchall()
    ]
    inv_name_entry["values"] = names

def add_inventory():

    name = inv_name_entry.get().strip().lower()
    if not name:
        messagebox.showerror("Error", "Enter ingredient name")
        return

    try:
        qty = float(inv_qty_entry.get())
        price = float(inv_price_entry.get())
    except:
        messagebox.showerror("Error", "Invalid quantity or price")
        return

    unit = inv_unit_entry.get().strip()

    added_value = qty * price

    c.execute("""
        SELECT qty, total_value FROM inventory
        WHERE LOWER(name)=?
    """, (name,))
    existing = c.fetchone()

    if existing:
        old_qty, old_value = existing

        new_qty = old_qty + qty
        new_value = old_value + added_value

        c.execute("""
            UPDATE inventory
            SET qty=?, total_value=?
            WHERE LOWER(name)=?
        """, (new_qty, new_value, name))

    else:
        c.execute("""
            INSERT INTO inventory(name, qty, unit, total_value)
            VALUES(?,?,?,?)
        """, (name, qty, unit, added_value))
    
    conn.commit()

    refresh_inventory()
    refresh_inventory_combo()
    refresh_recipe_combos()

    inv_qty_entry.delete(0, tk.END)
    inv_price_entry.delete(0, tk.END)

def get_selected_inventory_item():
    selected = inventory_tree.selection()
    if not selected:
        return None
    return inventory_tree.item(selected[0])["values"]

def delete_inventory_item():

    item = get_selected_inventory_item()
    if not item:
        messagebox.showerror("Error", "Select an ingredient to delete")
        return

    item_id, name = item[0], item[1]

    if not messagebox.askyesno(
        "Confirm Delete",
        f"Delete ingredient '{name}' completely?\n\n"
        "This will also remove it from all recipes.\n\n"
        "This action CANNOT be undone."
    ):
        return

    c.execute(
        "DELETE FROM recipes WHERE LOWER(ingredient)=?",
        (name.lower(),)
    )

    c.execute(
        "DELETE FROM inventory WHERE id=?",
        (item_id,)
    )

    conn.commit()

    refresh_inventory()
    refresh_inventory_combo()
    refresh_recipe_combos()

    inv_name_entry.delete(0, tk.END)
    inv_qty_entry.delete(0, tk.END)
    inv_unit_entry.delete(0, tk.END)
    inv_price_entry.delete(0, tk.END)

def on_inventory_select(event):
    name = inv_name_entry.get().strip().lower()
    c.execute(
        "SELECT unit FROM inventory WHERE LOWER(name)=?",
        (name,)
    )
    row = c.fetchone()
    if row and row[0]:
        inv_unit_entry.delete(0, tk.END)
        inv_unit_entry.insert(0, row[0])

inv_name_entry.bind("<<ComboboxSelected>>", on_inventory_select)    

def reset_inventory():

    if not messagebox.askyesno(
        "Confirm Reset",
        "This will reset ALL inventory quantities and values to ZERO.\n\nIngredients will NOT be deleted.\n\nContinue?"
    ):
        return

    c.execute("""
        UPDATE inventory
        SET qty = 0,
            total_value = 0
    """)

    conn.commit()

    inv_name_entry.delete(0, tk.END)
    inv_qty_entry.delete(0, tk.END)
    inv_unit_entry.delete(0, tk.END)
    inv_price_entry.delete(0, tk.END)

ttk.Button(
    inventory_form,
    text="Add Ingredient",
    command=add_inventory
).grid(row=4, column=0, columnspan=2, pady=(10, 5))

ttk.Button(
    inventory_form,
    text="🗑 Delete Ingredient",
    command=delete_inventory_item
).grid(row=5, column=0, columnspan=2, pady=(5, 10))

ttk.Button(
    inventory_frame,
    text="🔄 Reset",
    command=reset_inventory
).pack(pady=10)

# =========================================================
# PRODUCTS TAB
# =========================================================

products_frame = frames["Products"]

product_form = ttk.Frame(products_frame)
product_form.pack(pady=10)

ttk.Label(product_form, text="Product Name").grid(row=0, column=0, pady=5)
ttk.Label(product_form, text="Price").grid(row=1, column=0, pady=5)

product_name_entry = ttk.Entry(product_form, width=30)
product_price_entry = ttk.Entry(product_form, width=30)

product_name_entry.grid(row=0, column=1, pady=5)
product_price_entry.grid(row=1, column=1, pady=5)

product_tree = ttk.Treeview(
    products_frame,
    columns=("ID", "Name", "Price"),
    show="headings",
    height=20
)

for col in ("ID", "Name", "Price"):
    product_tree.heading(col, text=col)

product_tree.pack(fill="both", expand=True, padx=10, pady=10)


def refresh_products():

    for row in product_tree.get_children():
        product_tree.delete(row)

    c.execute("""
        SELECT id, name, price
        FROM products
        ORDER BY id
    """)

    rows = c.fetchall()

    for row in rows:
        product_tree.insert("", "end", values=row)

def add_product():

    name = product_name_entry.get().strip()

    if not name:
        messagebox.showerror("Error", "Enter product name")
        return

    try:
        price = float(product_price_entry.get())
    except:
        messagebox.showerror("Error", "Invalid price")
        return

    c.execute("""
        INSERT OR REPLACE INTO products(name, price)
        VALUES(?,?)
    """, (name, price))

    conn.commit()

    refresh_products()
    refresh_recipe_combos()
    refresh_sales_combo()

    product_name_entry.delete(0, tk.END)
    product_price_entry.delete(0, tk.END)

def delete_product():
    selected = product_tree.selection()
    if not selected:
        messagebox.showerror("Error", "Select a product to delete")
        return

    product_id, product_name, _ = product_tree.item(selected[0], "values")

    confirm = messagebox.askyesno(
        "Confirm Delete",
        f"Delete product '{product_name}'?\n\nRelated recipes will also be removed."
    )
    if not confirm:
        return

    # Remove related recipes first (data integrity)
    c.execute("DELETE FROM recipes WHERE product_id = ?", (product_id,))
    c.execute("DELETE FROM products WHERE id = ?", (product_id,))
    conn.commit()

    refresh_products()
    refresh_recipe_combos()
    refresh_sales_combo()

    product_name_entry.delete(0, tk.END)
    product_price_entry.delete(0, tk.END)

ttk.Button(
    product_form,
    text="Add Product",
    command=add_product
).grid(row=2, columnspan=2, pady=(10, 5))

ttk.Button(
    product_form,
    text="🗑 Delete Product",
    command=delete_product
).grid(row=3, columnspan=2, pady=(5, 10))

# =========================================================
# RECIPES TAB
# =========================================================

recipes_frame = frames["Recipes"]

recipe_form = ttk.Frame(recipes_frame)
recipe_form.pack(pady=10)

ttk.Label(recipe_form, text="Product").grid(row=0, column=0)
ttk.Label(recipe_form, text="Ingredient").grid(row=1, column=0)
ttk.Label(recipe_form, text="Qty Used").grid(row=2, column=0)

recipe_product_combo = ttk.Combobox(recipe_form, width=27)
recipe_ingredient_combo = ttk.Combobox(recipe_form, width=27)
recipe_qty_entry = ttk.Entry(recipe_form, width=30)

recipe_product_combo.grid(row=0, column=1, pady=5)
recipe_ingredient_combo.grid(row=1, column=1, pady=5)
recipe_qty_entry.grid(row=2, column=1, pady=5)

recipe_tree = ttk.Treeview(
    recipes_frame,
    columns=("ID", "Product", "Ingredient", "Qty"),
    show="headings",
    height=20
)

for col in ("ID", "Product", "Ingredient", "Qty"):
    recipe_tree.heading(col, text=col)

recipe_tree.pack(fill="both", expand=True, padx=10, pady=10)

def refresh_recipes():

    for row in recipe_tree.get_children():
        recipe_tree.delete(row)

    c.execute("""
        SELECT id, product, ingredient, qty
        FROM recipes
        ORDER BY product
    """)

    rows = c.fetchall()

    for row in rows:
        recipe_tree.insert("", "end", values=row)


def add_recipe():

    product = recipe_product_combo.get()
    ingredient = recipe_ingredient_combo.get()

    try:
        qty = float(recipe_qty_entry.get())
    except:
        messagebox.showerror("Error", "Invalid quantity")
        return

    c.execute("""
        INSERT INTO recipes(product, ingredient, qty)
        VALUES(?,?,?)
    """, (product, ingredient, qty))

    conn.commit()

    refresh_recipes()

    recipe_qty_entry.delete(0, tk.END)


ttk.Button(
    recipe_form,
    text="Add Recipe Item",
    command=add_recipe
).grid(row=3, columnspan=2, pady=10)

# =========================================================
# SALES TAB 
# =========================================================

sales_frame = frames["Sales"]
sales_form = ttk.Frame(sales_frame)
sales_form.pack(pady=10)

ttk.Label(sales_form, text="Product").grid(row=0, column=0)
ttk.Label(sales_form, text="Qty").grid(row=1, column=0)
ttk.Label(sales_form, text="Customer").grid(row=0, column=2)

sales_product_combo = ttk.Combobox(sales_form, width=30)
sales_product_combo.grid(row=0, column=1, pady=5)

sales_qty_entry = ttk.Entry(sales_form, width=30)
sales_qty_entry.grid(row=1, column=1, pady=5)

def toggle_pay_later(event=None):
    customer = sales_customer_combo.get().strip()

    if not customer:
        pay_later_var.set(False)          # Uncheck it
        pay_later_check.config(state="disabled") # Disable it
    else:
        pay_later_check.config(state="normal")   # Enable it

sales_customer_combo = ttk.Combobox(sales_form, width=25)
sales_customer_combo.grid(row=0, column=3, padx=5)
sales_customer_combo.bind("<<ComboboxSelected>>", toggle_pay_later)
sales_customer_combo.bind("<KeyRelease>", toggle_pay_later)


# ================= CUSTOMER SECTION =================

ttk.Label(sales_form, text="New Customer").grid(row=1, column=2)
new_customer_entry = ttk.Entry(sales_form, width=25)
new_customer_entry.grid(row=1, column=3)

ttk.Label(sales_form, text="Discount %").grid(row=2, column=2)
new_discount_entry = ttk.Entry(sales_form, width=25)
new_discount_entry.grid(row=2, column=3)

def load_customers():
    c.execute("SELECT name FROM customers")
    sales_customer_combo["values"] = [row[0] for row in c.fetchall()]

def add_customer():
    name = new_customer_entry.get().strip()

    if not name:
        messagebox.showerror("Error", "Enter customer name")
        return

    try:
        discount = float(new_discount_entry.get() or 0)
        if discount < 0:
            raise ValueError
    except:
        messagebox.showerror("Error", "Invalid discount")
        return

    try:
        c.execute("INSERT INTO customers (name, discount) VALUES (?, ?)", (name, discount))
        conn.commit()
        messagebox.showinfo("Success", "Customer added")
        new_customer_entry.delete(0, tk.END)
        new_discount_entry.delete(0, tk.END)
        load_customers()
    except sqlite3.IntegrityError:
        messagebox.showerror("Error", "Customer already exists")

ttk.Button(sales_form, text="Add Customer", command=add_customer)\
    .grid(row=3, column=3, pady=5)

load_customers()

# ================= CART =================

cart = []

cart_tree = ttk.Treeview(
    sales_frame,
    columns=("Product", "Qty", "Price", "Total"),
    show="headings",
    height=10
)

for col in ("Product", "Qty", "Price", "Total"):
    cart_tree.heading(col, text=col)

cart_tree.pack(fill="x", padx=10, pady=10)

cart_total_label = ttk.Label(
    sales_frame,
    text="TOTAL: 0.00 Toman",
    font=("Arial", 16, "bold")
)
cart_total_label.pack(pady=5)

def refresh_cart():
    cart_tree.delete(*cart_tree.get_children())

    total = 0
    for item in cart:
        cart_tree.insert("", "end", values=(
            item["product"],
            item["qty"],
            f"{item['price']:.2f}",
            f"{item['total']:.2f}"
        ))
        total += item["total"]

    cart_total_label.config(text=f"TOTAL: {total:.2f} Toman")

def add_to_cart():
    product = sales_product_combo.get().strip()

    if not product:
        messagebox.showerror("Error", "Select product")
        return

    try:
        qty = float(sales_qty_entry.get())
        if qty <= 0:
            raise ValueError
    except:
        messagebox.showerror("Error", "Invalid quantity")
        return

    c.execute("SELECT price FROM products WHERE name=?", (product,))
    result = c.fetchone()

    if not result:
        messagebox.showerror("Error", "Product not found")
        return

    price = result[0]

    # Merge if exists
    for item in cart:
        if item["product"] == product:
            item["qty"] += qty
            item["total"] = item["qty"] * item["price"]
            refresh_cart()
            sales_qty_entry.delete(0, tk.END)
            return

    cart.append({
        "product": product,
        "qty": qty,
        "price": price,
        "total": price * qty
    })

    refresh_cart()
    sales_qty_entry.delete(0, tk.END)

def remove_from_cart():
    selected = cart_tree.selection()
    if not selected:
        return

    index = cart_tree.index(selected[0])
    del cart[index]
    refresh_cart()

cart_tree.bind("<Double-1>", lambda e: remove_from_cart())

# ================= PAYMENT =================

pay_later_var = tk.BooleanVar()

pay_later_check = ttk.Checkbutton(
    sales_frame,
    text="Pay Later (Debt)",
    variable=pay_later_var
)
pay_later_check.pack(pady=5)
pay_later_check.config(state="disabled")

def print_receipt(customer, discount_pct, discount_val, total):
    """Generates a receipt and sends it to the default system printer."""
    receipt_text = "========== RECEIPT ==========\n"
    receipt_text += f"Customer: {customer or 'Walk-in'}\n"
    receipt_text += "-----------------------------\n"
    for item in cart:
        receipt_text += f"{item['product']} x {item['qty']} : ${item['total']:.2f}\n"
    receipt_text += "-----------------------------\n"
    receipt_text += f"Discount ({discount_pct}%): -${discount_val:.2f}\n"
    receipt_text += f"FINAL TOTAL: ${total:.2f}\n"
    receipt_text += "========== Thank You! =========="
    
    # 1. Save to a file first
    filename = "receipt_to_print.txt"
    with open(filename, "w") as f:
        f.write(receipt_text)
    
    # 2. Send to printer based on OS
    try:
        if platform.system() == "Windows":
            # Using 'print' command in Windows
            os.startfile(filename, "print")
        else:
            # For Linux/macOS
            os.system(f"lpr {filename}")
            
        messagebox.showinfo("Success", "Receipt sent to printer.")
    except Exception as e:
        messagebox.showerror("Print Error", f"Could not print: {str(e)}")

def checkout():
    """Processes sale, deducts inventory, and prints receipt."""
    try:
        if not cart:
            messagebox.showwarning("Error", "Cart is empty")
            return

        # 1. Gather info from UI
        customer_name = sales_customer_combo.get().strip()
        
        # Safely handle discount
        try:
            discount_percent = float(sales_discount_entry.get() or 0)
        except (NameError, ValueError, AttributeError):
            discount_percent = 0.0
            
        original_total = sum(item["total"] for item in cart)
        discount_amount = (original_total * discount_percent) / 100
        final_total = original_total - discount_amount
        payment_status = "Unpaid" if pay_later_var.get() else "Paid"

        # 2. Database Transaction
        conn.execute("BEGIN")
        
        try:
            for item in cart:
                # A. Deduct ingredients based on recipe
                c.execute("SELECT ingredient, qty FROM recipes WHERE product=?", (item["product"],))
                recipe_items = c.fetchall()

                for ingredient, recipe_qty in recipe_items:
                    total_reduction = recipe_qty * item["qty"]
                    c.execute("""
                        UPDATE inventory
                        SET qty = qty - ?
                        WHERE name = ?
                    """, (total_reduction, ingredient))

                # B. Insert sale record
                c.execute("""
                    INSERT INTO sales (product, qty, total, date, customer, payment_status)
                    VALUES (?, ?, ?, datetime('now','localtime'), ?, ?)
                """, (
                    item["product"],
                    item["qty"],
                    item["total"],
                    customer_name,
                    payment_status
                ))
            
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e

        # 3. Print receipt
        print_receipt(customer_name, discount_percent, discount_amount, final_total)

        # 4. Cleanup UI
        cart.clear()
        if 'refresh_cart' in globals(): refresh_cart()
        if 'refresh_inventory' in globals(): refresh_inventory()
        
        pay_later_var.set(False)
        sales_customer_combo.set("")
        if 'sales_discount_entry' in globals(): sales_discount_entry.delete(0, 'end')

        # 5. Refresh Debts
        if callable(globals().get("load_debts")): load_debts()

    except Exception as e:
        messagebox.showerror("Checkout Error", f"Failed to complete sale: {str(e)}")

     # 1. Clear the customer selection
    sales_customer_combo.set("") 
    
    # 2. Reset the pay_later checkbox
    pay_later_var.set(False) 
    
    # 3. Disable the checkbox immediately
    pay_later_check.config(state="disabled")

ttk.Button(sales_form, text="Add To Cart", command=add_to_cart)\
    .grid(row=2, column=0, pady=10)

ttk.Button(sales_form, text="Checkout", command=checkout)\
    .grid(row=2, column=1, pady=10)

ttk.Button(sales_form, text="Remove Item", command=remove_from_cart)\
    .grid(row=2, column=2, pady=10)

# =========================================================
# Sales History
# =========================================================

top_frame = ttk.Frame(history_tab)
top_frame.pack(fill="x", padx=10, pady=5)

total_label = ttk.Label(top_frame, text="Total Revenue: 0")
total_label.pack(side="right", padx=10)

refresh_button = ttk.Button(top_frame, text="Refresh")
refresh_button.pack(side="left", padx=5)

columns = ("ID", "Product", "Qty", "Total", "Date", "Customer", "Status")

history_tree = ttk.Treeview(
    history_tab,
    columns=columns,
    show="headings"
)

for col in columns:
    history_tree.heading(col, text=col)
    history_tree.column(col, anchor="center")

history_tree.pack(fill="both", expand=True, padx=10, pady=10)

# ---------------- REFRESH FUNCTION ----------------
def refresh_sales_history():
    for row in history_tree.get_children():
        history_tree.delete(row)

    c.execute("""
        SELECT id, product, qty, total, date, customer, payment_status
        FROM sales
        ORDER BY datetime(date) DESC
    """)

    rows = c.fetchall()

    total_revenue = 0

    for row in rows:
        history_tree.insert("", "end", values=row)
        total_revenue += row[3]  # total column

    total_label.config(text=f"Total Revenue: {total_revenue:.2f}")

refresh_button.config(command=refresh_sales_history)

refresh_sales_history()

# =========================================================
# DEBTS TAB
# =========================================================

debts_frame = frames["Debts"]

debts_tree = ttk.Treeview(
    debts_frame,
    columns=("ID", "Customer", "Product", "Amount", "Date"),
    show="headings",
    height=15
)

for col in ("ID", "Customer", "Product", "Amount", "Date"):
    debts_tree.heading(col, text=col)

debts_tree.pack(fill="both", expand=True, padx=10, pady=10)

def load_debts():

    debts_tree.delete(*debts_tree.get_children())

    c.execute("""
        SELECT id, customer, product, total, date
        FROM sales
        WHERE payment_status='Unpaid'
        ORDER BY datetime(date) DESC
    """)
    for row in c.fetchall():
        debts_tree.insert("", "end", values=row)

def mark_as_paid():
    selected = debts_tree.selection()
    if not selected:
        messagebox.showerror("Error", "Select a debt to mark as paid.")
        return

    sale_id = debts_tree.item(selected[0])["values"][0]

    c.execute("UPDATE sales SET payment_status='Paid' WHERE id=?", (sale_id,))
    conn.commit()

    messagebox.showinfo("Success", "Bill marked as paid.")

    load_debts()

    if callable(globals().get("generate_report")):
        generate_report()

ttk.Button(
    debts_frame,
    text="Mark Selected As Paid",
    command=mark_as_paid
).pack(pady=10)

ttk.Button(debts_frame, text="Refresh", command=load_debts).pack(pady=5)

# =========================================================
# REPORTS TAB
# =========================================================

reports_frame = frames["Reports"]

report_data = []

# ================= CONTROLS =================

report_controls = ttk.Frame(reports_frame)
report_controls.pack(fill="x", pady=10)

ttk.Label(report_controls, text="Report Type").pack(side="left", padx=5)

report_type_combo = ttk.Combobox(
    report_controls,
    values=["Daily", "Weekly", "Monthly", "Yearly"],
    state="readonly",
    width=20
)
report_type_combo.current(0)
report_type_combo.pack(side="left", padx=5)

report_text = tk.Text(
    reports_frame,
    font=("Consolas", 11)
)
report_text.pack(fill="both", expand=True, padx=10, pady=10)


# =========================================================
# GENERATE REPORT
# =========================================================
def generate_report():
    global report_data
    report_data = []

    report_text.delete("1.0", tk.END)
    report_type = report_type_combo.get()

    # ==========================================================
    # DATE FILTER
    # ==========================================================
    if report_type == "Daily":
        date_condition = "date(date) = date('now','localtime')"

    elif report_type == "Weekly":
        date_condition = "datetime(date) >= datetime('now','-7 days','localtime')"

    elif report_type == "Monthly":
        date_condition = """
            strftime('%Y-%m', datetime(date)) =
            strftime('%Y-%m', datetime('now','localtime'))
        """

    else:  # Yearly
        date_condition = """
            strftime('%Y', datetime(date)) =
            strftime('%Y', datetime('now','localtime'))
        """

    # ==========================================================
    # LOAD SALES
    # ==========================================================
    sales_query = f"""
        SELECT id, product, qty, total, date, customer, payment_status
        FROM sales
        WHERE {date_condition}
        ORDER BY datetime(date) DESC
    """
    c.execute(sales_query)
    sales_rows = c.fetchall()

    # ==========================================================
    # TITLE
    # ==========================================================
    report_text.insert(tk.END, "\n==============================\n")
    report_text.insert(tk.END, f"       {report_type.upper()} REPORT\n")
    report_text.insert(tk.END, "==============================\n\n")

    # ==========================================================
    # SUMMARY TOTALS
    # ==========================================================
    def calc_total(status):
        c.execute(f"""
            SELECT IFNULL(SUM(total), 0)
            FROM sales
            WHERE payment_status=? AND {date_condition}
        """, (status,))
        value = c.fetchone()[0]
        return value or 0

    paid_total = calc_total("Paid")
    unpaid_total = calc_total("Unpaid")
    grand_total = paid_total + unpaid_total

    c.execute(f"""
        SELECT COUNT(*)
        FROM sales
        WHERE {date_condition}
    """)
    sales_count = c.fetchone()[0] or 0

    report_text.insert(tk.END, "==== OVERVIEW ====\n")
    report_text.insert(tk.END, f"Sales Count    : {sales_count}\n")
    report_text.insert(tk.END, f"Paid Revenue   : {paid_total:.2f} Toman\n")
    report_text.insert(tk.END, f"Unpaid Debt    : {unpaid_total:.2f} Toman\n")
    report_text.insert(tk.END, f"Grand Total    : {grand_total:.2f} Toman\n\n")

    report_data.append({
        "Section": "Summary",
        "Sales Count": sales_count,
        "Paid Revenue": paid_total,
        "Unpaid Debt": unpaid_total,
        "Grand Total": grand_total
    })

    # ==========================================================
    # SALES DETAILS
    # ==========================================================
    report_text.insert(tk.END, "==== SALES DETAILS ====\n\n")

    if not sales_rows:
        report_text.insert(tk.END, "No sales found for this period.\n\n")
    else:
        for sale_id, product, sold_qty, total, sale_date, customer, payment_status in sales_rows:
            report_text.insert(
                tk.END,
                f"{sale_date} | {product} | Qty: {sold_qty} | Total: {total:.2f} | "
                f"Customer: {customer or '-'} | {payment_status}\n"
            )

            report_data.append({
                "Section": "Sales Details",
                "Sale ID": sale_id,
                "Date": sale_date,
                "Product": product,
                "Quantity": sold_qty,
                "Total": total,
                "Customer": customer or "",
                "Payment Status": payment_status
            })

    report_text.insert(tk.END, "\n")

    # ==========================================================
    # INGREDIENT / INVENTORY CONSUMPTION
    # ==========================================================
    report_text.insert(tk.END, "==== INVENTORY CONSUMPTION ====\n\n")

    ingredient_usage = {}

    for sale_id, product, sold_qty, total, sale_date, customer, payment_status in sales_rows:
        c.execute("SELECT ingredient, qty FROM recipes WHERE product=?", (product,))
        recipe_rows = c.fetchall()

        for ingredient, recipe_qty in recipe_rows:
            used = (recipe_qty or 0) * (sold_qty or 0)
            ingredient_usage[ingredient] = ingredient_usage.get(ingredient, 0) + used

    if not ingredient_usage:
        report_text.insert(tk.END, "No inventory consumption found.\n\n")
    else:
        for ingredient, used_qty in ingredient_usage.items():
            c.execute("SELECT qty, unit FROM inventory WHERE name=?", (ingredient,))
            inv = c.fetchone()
            remain, unit = (inv if inv else (0, ""))

            report_text.insert(tk.END, f"{ingredient.upper()}\n")
            report_text.insert(tk.END, f"   Consumed : {used_qty:.2f} {unit}\n")
            report_text.insert(tk.END, f"   Remaining: {remain:.2f} {unit}\n\n")

            report_data.append({
                "Section": "Inventory Consumption",
                "Ingredient": ingredient,
                "Consumed": used_qty,
                "Remaining": remain,
                "Unit": unit
            })

    # ==========================================================
    # CURRENT INVENTORY
    # ==========================================================
    report_text.insert(tk.END, "==== CURRENT INVENTORY ====\n\n")

    c.execute("SELECT name, qty, unit FROM inventory ORDER BY name")
    inventory_rows = c.fetchall()

    if not inventory_rows:
        report_text.insert(tk.END, "No inventory found.\n\n")
    else:
        for name, qty, unit in inventory_rows:
            report_text.insert(tk.END, f"{name}: {qty:.2f} {unit}\n")

            report_data.append({
                "Section": "Current Inventory",
                "Item": name,
                "Quantity": qty,
                "Unit": unit
            })

    report_text.insert(tk.END, "\n")

    # ==========================================================
    # CUSTOMER SUMMARY
    # ==========================================================
    report_text.insert(tk.END, "==== CUSTOMER SUMMARY ====\n\n")

    c.execute(f"""
        SELECT 
            customer,
            COUNT(id),
            IFNULL(SUM(CASE WHEN payment_status='Paid' THEN total ELSE 0 END), 0),
            IFNULL(SUM(CASE WHEN payment_status='Unpaid' THEN total ELSE 0 END), 0)
        FROM sales
        WHERE customer IS NOT NULL
          AND customer != ''
          AND {date_condition}
        GROUP BY customer
        ORDER BY customer
    """)
    customer_rows = c.fetchall()

    total_customer_debt = 0

    if not customer_rows:
        report_text.insert(tk.END, "No customer data for this period.\n\n")
    else:
        for customer, sale_count, paid, unpaid in customer_rows:
            paid = paid or 0
            unpaid = unpaid or 0
            total_customer_debt += unpaid

            report_text.insert(tk.END, f"{customer}\n")
            report_text.insert(tk.END, f"   Purchases : {sale_count}\n")
            report_text.insert(tk.END, f"   Paid      : {paid:.2f} Toman\n")
            report_text.insert(tk.END, f"   Debt      : {unpaid:.2f} Toman\n\n")

            report_data.append({
                "Section": "Customer Summary",
                "Customer": customer,
                "Purchases": sale_count,
                "Paid": paid,
                "Debt": unpaid
            })

    report_text.insert(tk.END, f"TOTAL CUSTOMER DEBT: {total_customer_debt:.2f} Toman\n\n")

    report_data.append({
        "Section": "Customer Debt Total",
        "Total Customer Debt": total_customer_debt
    })

    # ==========================================================
    # UNPAID BILLS DETAILS
    # ==========================================================
    report_text.insert(tk.END, "==== UNPAID BILLS ====\n\n")

    c.execute(f"""
        SELECT id, customer, product, qty, total, date
        FROM sales
        WHERE payment_status='Unpaid' AND {date_condition}
        ORDER BY datetime(date) DESC
    """)
    unpaid_rows = c.fetchall()

    if not unpaid_rows:
        report_text.insert(tk.END, "No unpaid bills for this period.\n")
    else:
        for sale_id, customer, product, qty, total, sale_date in unpaid_rows:
            report_text.insert(
                tk.END,
                f"ID: {sale_id} | {sale_date} | Customer: {customer or '-'} | "
                f"{product} | Qty: {qty} | Debt: {total:.2f} Toman\n"
            )

            report_data.append({
                "Section": "Unpaid Bills",
                "Sale ID": sale_id,
                "Date": sale_date,
                "Customer": customer or "",
                "Product": product,
                "Quantity": qty,
                "Debt": total
            })


# =========================================================
# EXPORT TO EXCEL
# =========================================================

def export_report_excel():
    report_type = report_type_combo.get()

    if report_type == "Daily":
        date_condition = "date(date) = date('now','localtime')"
    elif report_type == "Weekly":
        date_condition = "datetime(date) >= datetime('now','-7 days','localtime')"
    elif report_type == "Monthly":
        date_condition = """
            strftime('%Y-%m', datetime(date)) =
            strftime('%Y-%m', datetime('now','localtime'))
        """
    else:  # Yearly
        date_condition = """
            strftime('%Y', datetime(date)) =
            strftime('%Y', datetime('now','localtime'))
        """

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")],
        title="Save Full Report As"
    )

    if not file_path:
        return

    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.styles import Font

        # =====================================================
        # 1) SUMMARY
        # =====================================================
        c.execute(f"""
            SELECT COUNT(*),
                   IFNULL(SUM(CASE WHEN payment_status='Paid' THEN total ELSE 0 END), 0),
                   IFNULL(SUM(CASE WHEN payment_status='Unpaid' THEN total ELSE 0 END), 0)
            FROM sales
            WHERE {date_condition}
        """)
        sales_count, paid_total, unpaid_total = c.fetchone()
        grand_total = (paid_total or 0) + (unpaid_total or 0)

        summary_df = pd.DataFrame([
            ["Sales Count", sales_count],
            ["Paid Revenue", paid_total],
            ["Unpaid Debt", unpaid_total],
            ["Grand Total", grand_total],
        ], columns=["Metric", "Value"])

        # =====================================================
        # 2) SALES DETAILS
        # =====================================================
        c.execute(f"""
            SELECT id, product, qty, total, date, customer, payment_status
            FROM sales
            WHERE {date_condition}
            ORDER BY datetime(date) DESC
        """)
        sales_rows = c.fetchall()
        sales_df = pd.DataFrame(
            sales_rows,
            columns=["Sale ID", "Product", "Quantity", "Total", "Date", "Customer", "Payment Status"]
        )

        # =====================================================
        # 3) INVENTORY CONSUMPTION
        # =====================================================
        ingredient_usage = {}

        for sale_id, product, sold_qty, total, sale_date, customer, payment_status in sales_rows:
            c.execute("SELECT ingredient, qty FROM recipes WHERE product=?", (product,))
            for ingredient, recipe_qty in c.fetchall():
                used = (recipe_qty or 0) * (sold_qty or 0)
                ingredient_usage[ingredient] = ingredient_usage.get(ingredient, 0) + used

        consumption_rows = []
        for ingredient, used_qty in ingredient_usage.items():
            c.execute("SELECT qty, unit FROM inventory WHERE name=?", (ingredient,))
            inv = c.fetchone()
            remain, unit = inv if inv else (0, "")
            consumption_rows.append([ingredient, used_qty, remain, unit])

        consumption_df = pd.DataFrame(
            consumption_rows,
            columns=["Ingredient", "Consumed", "Remaining", "Unit"]
        )

        # =====================================================
        # 4) CURRENT INVENTORY
        # =====================================================
        c.execute("SELECT name, qty, unit FROM inventory ORDER BY name")
        inventory_rows = c.fetchall()
        inventory_df = pd.DataFrame(
            inventory_rows,
            columns=["Item", "Quantity", "Unit"]
        )

        # =====================================================
        # 5) CUSTOMER SUMMARY
        # =====================================================
        c.execute(f"""
            SELECT 
                customer,
                COUNT(id),
                IFNULL(SUM(CASE WHEN payment_status='Paid' THEN total ELSE 0 END), 0),
                IFNULL(SUM(CASE WHEN payment_status='Unpaid' THEN total ELSE 0 END), 0)
            FROM sales
            WHERE customer IS NOT NULL
              AND customer != ''
              AND {date_condition}
            GROUP BY customer
            ORDER BY customer
        """)
        customer_rows = c.fetchall()
        customer_df = pd.DataFrame(
            customer_rows,
            columns=["Customer", "Purchases", "Paid", "Debt"]
        )

        # =====================================================
        # 6) UNPAID BILLS
        # =====================================================
        c.execute(f"""
            SELECT id, customer, product, qty, total, date
            FROM sales
            WHERE payment_status='Unpaid' AND {date_condition}
            ORDER BY datetime(date) DESC
        """)
        unpaid_rows = c.fetchall()
        unpaid_df = pd.DataFrame(
            unpaid_rows,
            columns=["Sale ID", "Customer", "Product", "Quantity", "Debt", "Date"]
        )

        # =====================================================
        # WRITE TO EXCEL - SEPARATE SHEETS
        # =====================================================
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            sales_df.to_excel(writer, sheet_name="Sales Details", index=False)
            consumption_df.to_excel(writer, sheet_name="Inventory Consumption", index=False)
            inventory_df.to_excel(writer, sheet_name="Current Inventory", index=False)
            customer_df.to_excel(writer, sheet_name="Customer Summary", index=False)
            unpaid_df.to_excel(writer, sheet_name="Unpaid Bills", index=False)

        # =====================================================
        # ADD CHARTS WITH OPENPYXL
        # =====================================================
        wb = load_workbook(file_path)

        # ---------- Summary chart ----------
        ws = wb["Summary"]
        ws["D1"] = "Charts"
        ws["D1"].font = Font(bold=True)

        chart1 = BarChart()
        chart1.title = "Financial Overview"
        chart1.y_axis.title = "Amount"
        chart1.x_axis.title = "Metric"

        data = Reference(ws, min_col=2, min_row=2, max_row=4)
        cats = Reference(ws, min_col=1, min_row=2, max_row=4)
        chart1.add_data(data, titles_from_data=False)
        chart1.set_categories(cats)
        chart1.height = 7
        chart1.width = 12
        ws.add_chart(chart1, "D3")

        # ---------- Sales Details chart ----------
        if "Sales Details" in wb.sheetnames and len(sales_df) > 0:
            ws = wb["Sales Details"]
            sales_product_summary = sales_df.groupby("Product", as_index=False)["Quantity"].sum()

            temp_row = len(sales_df) + 3
            ws.cell(row=temp_row, column=1, value="Product")
            ws.cell(row=temp_row, column=2, value="Total Qty")

            for i, row in enumerate(sales_product_summary.values.tolist(), start=temp_row + 1):
                ws.cell(row=i, column=1, value=row[0])
                ws.cell(row=i, column=2, value=row[1])

            chart2 = BarChart()
            chart2.title = "Product Sales Quantity"
            chart2.y_axis.title = "Quantity Sold"
            chart2.x_axis.title = "Product"

            data = Reference(ws, min_col=2, min_row=temp_row, max_row=i)
            cats = Reference(ws, min_col=1, min_row=temp_row + 1, max_row=i)
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(cats)
            chart2.height = 8
            chart2.width = 14
            ws.add_chart(chart2, "I2")

        # ---------- Inventory Consumption chart ----------
        if "Inventory Consumption" in wb.sheetnames and len(consumption_df) > 0:
            ws = wb["Inventory Consumption"]

            chart3 = BarChart()
            chart3.title = "Inventory Consumption"
            chart3.y_axis.title = "Consumed"
            chart3.x_axis.title = "Ingredient"

            data = Reference(ws, min_col=2, min_row=1, max_row=len(consumption_df) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(consumption_df) + 1)
            chart3.add_data(data, titles_from_data=True)
            chart3.set_categories(cats)
            chart3.height = 8
            chart3.width = 14
            ws.add_chart(chart3, "F2")

        # ---------- Current Inventory chart ----------
        if "Current Inventory" in wb.sheetnames and len(inventory_df) > 0:
            ws = wb["Current Inventory"]

            chart4 = BarChart()
            chart4.title = "Current Inventory Levels"
            chart4.y_axis.title = "Quantity"
            chart4.x_axis.title = "Item"

            data = Reference(ws, min_col=2, min_row=1, max_row=len(inventory_df) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(inventory_df) + 1)
            chart4.add_data(data, titles_from_data=True)
            chart4.set_categories(cats)
            chart4.height = 8
            chart4.width = 14
            ws.add_chart(chart4, "E2")

        # ---------- Customer Summary chart ----------
        if "Customer Summary" in wb.sheetnames and len(customer_df) > 0:
            ws = wb["Customer Summary"]

            chart5 = BarChart()
            chart5.title = "Customer Debt"
            chart5.y_axis.title = "Debt"
            chart5.x_axis.title = "Customer"

            data = Reference(ws, min_col=4, min_row=1, max_row=len(customer_df) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(customer_df) + 1)
            chart5.add_data(data, titles_from_data=True)
            chart5.set_categories(cats)
            chart5.height = 8
            chart5.width = 14
            ws.add_chart(chart5, "F2")

            # Optional pie chart for Paid vs Debt of all customers
            total_paid = sum(customer_df["Paid"]) if len(customer_df) > 0 else 0
            total_debt = sum(customer_df["Debt"]) if len(customer_df) > 0 else 0

            ws["H20"] = "Type"
            ws["I20"] = "Amount"
            ws["H21"] = "Paid"
            ws["I21"] = total_paid
            ws["H22"] = "Debt"
            ws["I22"] = total_debt

            pie = PieChart()
            pie.title = "Paid vs Debt"
            labels = Reference(ws, min_col=8, min_row=21, max_row=22)
            data = Reference(ws, min_col=9, min_row=20, max_row=22)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 8
            pie.width = 10
            ws.add_chart(pie, "H2")

        # ---------- Unpaid Bills chart ----------
        if "Unpaid Bills" in wb.sheetnames and len(unpaid_df) > 0:
            ws = wb["Unpaid Bills"]

            unpaid_customer_summary = unpaid_df.groupby("Customer", as_index=False)["Debt"].sum()

            temp_row = len(unpaid_df) + 3
            ws.cell(row=temp_row, column=1, value="Customer")
            ws.cell(row=temp_row, column=2, value="Debt")

            for i, row in enumerate(unpaid_customer_summary.values.tolist(), start=temp_row + 1):
                ws.cell(row=i, column=1, value=row[0])
                ws.cell(row=i, column=2, value=row[1])

            chart6 = BarChart()
            chart6.title = "Unpaid Debt by Customer"
            chart6.y_axis.title = "Debt"
            chart6.x_axis.title = "Customer"

            data = Reference(ws, min_col=2, min_row=temp_row, max_row=i)
            cats = Reference(ws, min_col=1, min_row=temp_row + 1, max_row=i)
            chart6.add_data(data, titles_from_data=True)
            chart6.set_categories(cats)
            chart6.height = 8
            chart6.width = 14
            ws.add_chart(chart6, "H2")

        wb.save(file_path)

        messagebox.showinfo("Success", f"Report exported successfully:\n{file_path}")

    except Exception as e:
        messagebox.showerror("Export Error", str(e))

ttk.Button(
    report_controls,
    text="Generate Report",
    command=generate_report
).pack(side="left", padx=10)

ttk.Button(
    report_controls,
    text="Export Excel",
    command=export_report_excel
).pack(side="left", padx=10)


# =========================================================
# INITIAL LOAD
# =========================================================
refresh_inventory()
refresh_products()
refresh_recipes()
refresh_recipe_combos()
refresh_sales_combo()
refresh_inventory_combo()

if "load_debts" in globals():
    load_debts()


# =========================================================
# LOGIN WINDOW
# =========================================================
def show_login():
    login = tk.Toplevel(root)
    login.title("Login")
    login.grab_set()
    login.transient(root)

    ttk.Label(login, text="Username").grid(row=0, column=0, pady=5, padx=5)
    ttk.Label(login, text="Password").grid(row=1, column=0, pady=5, padx=5)

    username_entry = ttk.Entry(login)
    password_entry = ttk.Entry(login, show="*")

    username_entry.grid(row=0, column=1, pady=5, padx=5)
    password_entry.grid(row=1, column=1, pady=5, padx=5)

    def do_login():
        username = username_entry.get()
        password = password_entry.get()

        c.execute(
            "SELECT role FROM users WHERE username=? AND password=?",
            (username, password)
        )
        result = c.fetchone()

        if not result:
            messagebox.showerror("Login Failed", "Invalid credentials")
            return

        root.current_role = result[0]
        login.destroy()
        apply_permissions()

    ttk.Button(login, text="Login", command=do_login).grid(
        row=2, column=0, columnspan=2, pady=10
    )

    username_entry.focus_set()

    username_entry.bind("<Return>", lambda e: password_entry.focus_set())
    password_entry.bind("<Return>", lambda e: do_login())
    login.bind("<Return>", lambda e: do_login())


# =========================================================
# APPLY PERMISSIONS TO TABS
# =========================================================
def apply_permissions():
    allowed = ROLE_PERMISSIONS.get(root.current_role, [])

    for tab_name, frame in frames.items():
        tab_index = notebook.index(frame)
        state = "normal" if tab_name in allowed else "disabled"
        notebook.tab(tab_index, state=state)

# =========================================================
# USERS & PERMISSION SYSTEM
# =========================================================

# ---- ROLE DEFINITIONS ----
ROLE_PERMISSIONS = {
    "admin":    ["Inventory", "Products", "Recipes", "Sales", "Reports", "Debts", "Users"],
    "manager":  ["Inventory", "Products", "Recipes", "Sales", "Reports", "Debts", "Users"],
    "salesman": ["Sales", "Reports", "Debts"],
    "stockman": ["Inventory"]
}

# ---- CREATE USERS TABLE ----
c.execute("""
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE NOT NULL,
    password TEXT NOT NULL,
    role TEXT NOT NULL
)
""")
conn.commit()

# ---- ENSURE DEFAULT ADMIN EXISTS ----
c.execute("SELECT * FROM users WHERE username='admin'")
if not c.fetchone():
    c.execute(
        "INSERT INTO users (username, password, role) VALUES (?,?,?)",
        ("admin", "admin", "admin")
    )
    conn.commit()


# ============================================================
# USERS TAB 
# ============================================================

users_frame = ttk.Frame(notebook)
notebook.add(users_frame, text="Users")
frames["Users"] = users_frame


# ---- FUNCTIONS ----
def refresh_users():
    users_tree.delete(*users_tree.get_children())
    c.execute("SELECT id, username, role FROM users")
    for row in c.fetchall():
        users_tree.insert("", "end", values=row)


def add_or_update_user():
    username = user_entry.get().strip()
    password = pass_entry.get().strip()
    role = role_combo.get()

    if not username or not role:
        messagebox.showwarning("Error", "Username and role are required")
        return

    selected = users_tree.selection()

    try:
        if selected:
            # UPDATE USER
            user_id = users_tree.item(selected[0])["values"][0]

            if password:
                c.execute(
                    "UPDATE users SET username=?, password=?, role=? WHERE id=?",
                    (username, password, role, user_id)
                )
            else:
                c.execute(
                    "UPDATE users SET username=?, role=? WHERE id=?",
                    (username, role, user_id)
                )
        else:
            # ADD NEW USER
            if not password:
                messagebox.showwarning("Error", "Password required for new user")
                return

            c.execute(
                "INSERT INTO users (username, password, role) VALUES (?,?,?)",
                (username, password, role)
            )

        conn.commit()
        refresh_users()
        user_entry.delete(0, "end")
        pass_entry.delete(0, "end")

    except sqlite3.IntegrityError:
        messagebox.showerror("Error", "Username already exists")

def switch_user():
    root.current_role = None

    # Disable all tabs until login
    for tab_name, frame in frames.items():
        index = notebook.index(frame)
        notebook.tab(index, state="disabled")

    show_login()

def delete_user():
    selected = users_tree.selection()
    if not selected:
        return

    user_id, username, _ = users_tree.item(selected[0])["values"]

    if username == "admin":
        messagebox.showerror("Error", "Cannot delete the admin user")
        return

    if messagebox.askyesno("Confirm", f"Delete user '{username}'?"):
        c.execute("DELETE FROM users WHERE id=?", (user_id,))
        conn.commit()
        refresh_users()


def on_user_select(event):
    selected = users_tree.selection()
    if not selected:
        return

    _, username, role = users_tree.item(selected[0])["values"]
    user_entry.delete(0, "end")
    pass_entry.delete(0, "end")
    user_entry.insert(0, username)
    role_combo.set(role)

# ---- USERS TAB UI ----
form = ttk.Frame(users_frame)
form.pack(pady=10)

ttk.Label(form, text="Username").grid(row=0, column=0, padx=5)
ttk.Label(form, text="Password").grid(row=1, column=0, padx=5)
ttk.Label(form, text="Role").grid(row=2, column=0, padx=5)

user_entry = ttk.Entry(form)
pass_entry = ttk.Entry(form, show="*")
role_combo = ttk.Combobox(
    form,
    values=["manager", "salesman", "stockman"],
    state="readonly"
)

user_entry.grid(row=0, column=1)
pass_entry.grid(row=1, column=1)
role_combo.grid(row=2, column=1)

ttk.Button(form, text="Add / Update User", command=add_or_update_user).grid(
    row=3, columnspan=2, pady=5
)
ttk.Button(form, text="Delete User", command=delete_user).grid(
    row=4, columnspan=2
)

users_tree = ttk.Treeview(
    users_frame,
    columns=("ID", "Username", "Role"),
    show="headings"
)
users_tree.heading("ID", text="ID")
users_tree.heading("Username", text="Username")
users_tree.heading("Role", text="Role")
users_tree.pack(fill="both", expand=True, padx=10, pady=10)

users_tree.bind("<<TreeviewSelect>>", on_user_select)

refresh_users()

# ============================================================
#                     LOGIN SYSTEM
# ============================================================

login_attempts = 0
login_window = None

def show_login():
    global login_window, login_attempts

    if login_window and login_window.winfo_exists():
        return

    login_window = tk.Toplevel(root)
    login_window.title("Login")
    login_window.grab_set()
    login_window.protocol("WM_DELETE_WINDOW", root.destroy)

    ttk.Label(login_window, text="Username").grid(row=0, column=0, padx=5, pady=5)
    ttk.Label(login_window, text="Password").grid(row=1, column=0, padx=5, pady=5)

    # Load usernames
    c.execute("SELECT username FROM users")
    usernames = [u[0] for u in c.fetchall()]

    username_combo = ttk.Combobox(login_window, values=usernames, state="readonly")
    username_combo.set("Select User")
    username_combo.grid(row=0, column=1, padx=5)

    password_entry = ttk.Entry(login_window, show="*")
    password_entry.grid(row=1, column=1, padx=5)

    def do_login():
        nonlocal username_combo, password_entry
        global login_attempts

        username = username_combo.get()
        password = password_entry.get()

        if username == "Select User":
            messagebox.showwarning("Warning", "Please select a user")
            return

        c.execute(
            "SELECT role FROM users WHERE username=? AND password=?",
            (username, password)
        )
        row = c.fetchone()

        if not row:
            login_attempts += 1
            messagebox.showerror("Login Failed", f"Incorrect password. ({login_attempts}/3)")
            if login_attempts >= 3:
                root.destroy()
            return

        root.current_role = row[0]
        login_window.destroy()
        apply_permissions()

    ttk.Button(login_window, text="Login", command=do_login).grid(
        row=2, columnspan=2, pady=10
    )

# OPEN LOGIN AT START
root.current_role = None
root.after(200, show_login)

# ---- TOP BAR ----
top_bar = ttk.Frame(root)
top_bar.pack(fill="x")

ttk.Button(top_bar, text="Switch User", command=switch_user).pack(
    side="right", padx=10, pady=5
)

# ================= END USERS SYSTEM =================

root.mainloop()