"""Excel export: build an in-memory .xlsx from a report dict (port of legacy Sante.py)."""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font


def build_excel(report: dict) -> bytes:
    """Return raw bytes of an openpyxl workbook for the given report dict."""

    overview = report["overview"]
    sales_details = report["sales_details"]
    inventory_consumption = report["inventory_consumption"]
    current_inventory = report["current_inventory"]
    customer_summary = report["customer_summary"]
    unpaid_bills = report["unpaid_bills"]

    # ----------------------------------------------------------------
    # Build DataFrames
    # ----------------------------------------------------------------
    summary_df = pd.DataFrame(
        [
            ["Sales Count", overview["sales_count"]],
            ["Paid Revenue", overview["paid_revenue"]],
            ["Unpaid Debt", overview["unpaid_debt"]],
            ["Grand Total", overview["grand_total"]],
        ],
        columns=["Metric", "Value"],
    )

    if sales_details:
        sales_df = pd.DataFrame(
            [
                [
                    r["sale_id"],
                    r["product"],
                    r["qty"],
                    r["line_total"],
                    r["date"],
                    r["customer"],
                    r["payment_status"],
                ]
                for r in sales_details
            ],
            columns=["Sale ID", "Product", "Quantity", "Line Total", "Date", "Customer", "Payment Status"],
        )
    else:
        sales_df = pd.DataFrame(
            columns=["Sale ID", "Product", "Quantity", "Line Total", "Date", "Customer", "Payment Status"]
        )

    if inventory_consumption:
        consumption_df = pd.DataFrame(
            [[r["ingredient"], r["consumed"], r["remaining"], r["unit"]] for r in inventory_consumption],
            columns=["Ingredient", "Consumed", "Remaining", "Unit"],
        )
    else:
        consumption_df = pd.DataFrame(columns=["Ingredient", "Consumed", "Remaining", "Unit"])

    if current_inventory:
        inventory_df = pd.DataFrame(
            [[r["name"], r["qty"], r["unit"]] for r in current_inventory],
            columns=["Item", "Quantity", "Unit"],
        )
    else:
        inventory_df = pd.DataFrame(columns=["Item", "Quantity", "Unit"])

    if customer_summary:
        customer_df = pd.DataFrame(
            [[r["customer"], r["purchases"], r["paid"], r["debt"]] for r in customer_summary],
            columns=["Customer", "Purchases", "Paid", "Debt"],
        )
    else:
        customer_df = pd.DataFrame(columns=["Customer", "Purchases", "Paid", "Debt"])

    if unpaid_bills:
        unpaid_df = pd.DataFrame(
            [[r["sale_id"], r["customer"], r["total"], r["date"]] for r in unpaid_bills],
            columns=["Sale ID", "Customer", "Debt", "Date"],
        )
    else:
        unpaid_df = pd.DataFrame(columns=["Sale ID", "Customer", "Debt", "Date"])

    # ----------------------------------------------------------------
    # Write sheets to an in-memory buffer with pandas
    # ----------------------------------------------------------------
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        sales_df.to_excel(writer, sheet_name="Sales Details", index=False)
        consumption_df.to_excel(writer, sheet_name="Inventory Consumption", index=False)
        inventory_df.to_excel(writer, sheet_name="Current Inventory", index=False)
        customer_df.to_excel(writer, sheet_name="Customer Summary", index=False)
        unpaid_df.to_excel(writer, sheet_name="Unpaid Bills", index=False)

    buf.seek(0)

    # ----------------------------------------------------------------
    # Add charts with openpyxl
    # ----------------------------------------------------------------
    wb = load_workbook(buf)

    # ---------- Summary chart (bar): Paid Revenue / Unpaid Debt / Grand Total ----------
    ws = wb["Summary"]
    ws["D1"] = "Charts"
    ws["D1"].font = Font(bold=True)

    chart1 = BarChart()
    chart1.title = "Financial Overview"
    chart1.y_axis.title = "Amount"
    chart1.x_axis.title = "Metric"

    data = Reference(ws, min_col=2, min_row=2, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.height = 7
    chart1.width = 12
    ws.add_chart(chart1, "D3")

    # ---------- Sales Details chart: product quantity bar ----------
    if len(sales_df) > 0:
        ws = wb["Sales Details"]
        sales_product_summary = sales_df.groupby("Product", as_index=False)["Quantity"].sum()

        temp_row = len(sales_df) + 3
        ws.cell(row=temp_row, column=1, value="Product")
        ws.cell(row=temp_row, column=2, value="Total Qty")

        last_i = temp_row
        for i, row in enumerate(sales_product_summary.values.tolist(), start=temp_row + 1):
            ws.cell(row=i, column=1, value=row[0])
            ws.cell(row=i, column=2, value=row[1])
            last_i = i

        chart2 = BarChart()
        chart2.title = "Product Sales Quantity"
        chart2.y_axis.title = "Quantity Sold"
        chart2.x_axis.title = "Product"

        data = Reference(ws, min_col=2, min_row=temp_row, max_row=last_i)
        cats = Reference(ws, min_col=1, min_row=temp_row + 1, max_row=last_i)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.height = 8
        chart2.width = 14
        ws.add_chart(chart2, "I2")

    # ---------- Inventory Consumption chart ----------
    if len(consumption_df) > 0:
        ws = wb["Inventory Consumption"]

        chart3 = BarChart()
        chart3.title = "Inventory Consumption"
        chart3.y_axis.title = "Consumed"
        chart3.x_axis.title = "Ingredient"

        data = Reference(ws, min_col=2, min_row=1, max_row=len(consumption_df) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(consumption_df) + 1)
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(cats)
        chart3.height = 8
        chart3.width = 14
        ws.add_chart(chart3, "F2")

    # ---------- Current Inventory chart ----------
    if len(inventory_df) > 0:
        ws = wb["Current Inventory"]

        chart4 = BarChart()
        chart4.title = "Current Inventory Levels"
        chart4.y_axis.title = "Quantity"
        chart4.x_axis.title = "Item"

        data = Reference(ws, min_col=2, min_row=1, max_row=len(inventory_df) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(inventory_df) + 1)
        chart4.add_data(data, titles_from_data=True)
        chart4.set_categories(cats)
        chart4.height = 8
        chart4.width = 14
        ws.add_chart(chart4, "E2")

    # ---------- Customer Summary chart ----------
    if len(customer_df) > 0:
        ws = wb["Customer Summary"]

        chart5 = BarChart()
        chart5.title = "Customer Debt"
        chart5.y_axis.title = "Debt"
        chart5.x_axis.title = "Customer"

        data = Reference(ws, min_col=4, min_row=1, max_row=len(customer_df) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(customer_df) + 1)
        chart5.add_data(data, titles_from_data=True)
        chart5.set_categories(cats)
        chart5.height = 8
        chart5.width = 14
        ws.add_chart(chart5, "F2")

        # Pie chart: Paid vs Debt
        total_paid = float(customer_df["Paid"].sum())
        total_debt = float(customer_df["Debt"].sum())

        ws["H20"] = "Type"
        ws["I20"] = "Amount"
        ws["H21"] = "Paid"
        ws["I21"] = total_paid
        ws["H22"] = "Debt"
        ws["I22"] = total_debt

        pie = PieChart()
        pie.title = "Paid vs Debt"
        labels = Reference(ws, min_col=8, min_row=21, max_row=22)
        pie_data = Reference(ws, min_col=9, min_row=20, max_row=22)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 8
        pie.width = 10
        ws.add_chart(pie, "H2")

    # ---------- Unpaid Bills chart ----------
    if len(unpaid_df) > 0:
        ws = wb["Unpaid Bills"]

        unpaid_customer_summary = unpaid_df.groupby("Customer", as_index=False)["Debt"].sum()

        temp_row = len(unpaid_df) + 3
        ws.cell(row=temp_row, column=1, value="Customer")
        ws.cell(row=temp_row, column=2, value="Debt")

        last_i = temp_row
        for i, row in enumerate(unpaid_customer_summary.values.tolist(), start=temp_row + 1):
            ws.cell(row=i, column=1, value=row[0])
            ws.cell(row=i, column=2, value=row[1])
            last_i = i

        chart6 = BarChart()
        chart6.title = "Unpaid Debt by Customer"
        chart6.y_axis.title = "Debt"
        chart6.x_axis.title = "Customer"

        data = Reference(ws, min_col=2, min_row=temp_row, max_row=last_i)
        cats = Reference(ws, min_col=1, min_row=temp_row + 1, max_row=last_i)
        chart6.add_data(data, titles_from_data=True)
        chart6.set_categories(cats)
        chart6.height = 8
        chart6.width = 14
        ws.add_chart(chart6, "H2")

    # ----------------------------------------------------------------
    # Save to buffer and return bytes
    # ----------------------------------------------------------------
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
