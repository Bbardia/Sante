"""Tests for /reports JSON and /reports/export.xlsx endpoints."""

from io import BytesIO

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _login(client, username="admin", password="admin"):
    resp = client.post("/auth/login", json={"username": username, "password": password})
    assert resp.status_code == 200, resp.text
    return resp.json()["access_token"]


def _auth(token):
    return {"Authorization": f"Bearer {token}"}


def _create_user(client, token, username, role):
    resp = client.post(
        "/users",
        json={"username": username, "password": "pass", "role": role},
        headers=_auth(token),
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


def _setup_world(client, token):
    """
    Set up:
      - ingredient: flour (qty=200, unit=kg, price=2)
      - product: bread (price=5)
      - recipe: bread <- flour qty=2
      - customer: Alice

    Returns (flour_id, bread_id, alice_id).
    """
    flour = client.post(
        "/inventory",
        json={"name": "flour", "qty": 200.0, "unit": "kg", "price": 2.0},
        headers=_auth(token),
    ).json()
    bread = client.post(
        "/products",
        json={"name": "bread", "price": 5.0},
        headers=_auth(token),
    ).json()
    client.post(
        "/recipes",
        json={"product_id": bread["id"], "ingredient_id": flour["id"], "qty": 2.0},
        headers=_auth(token),
    )
    alice = client.post(
        "/customers",
        json={"name": "Alice"},
        headers=_auth(token),
    ).json()
    return flour["id"], bread["id"], alice["id"]


def _sell(client, token, bread_id, qty, alice_id=None, pay_later=False):
    body = {"items": [{"product_id": bread_id, "qty": qty}]}
    if alice_id is not None:
        body["customer_id"] = alice_id
    if pay_later:
        body["pay_later"] = True
    resp = client.post("/sales", json=body, headers=_auth(token))
    assert resp.status_code == 200, resp.text
    return resp.json()


# ---------------------------------------------------------------------------
# Overview & basic report structure
# ---------------------------------------------------------------------------

def test_report_overview_correct_for_today(client):
    """Default Daily report reflects today's sales correctly."""
    token = _login(client)
    flour_id, bread_id, alice_id = _setup_world(client, token)

    # Paid sale: 2 breads = 10.0
    _sell(client, token, bread_id, qty=2)
    # Unpaid sale for Alice: 3 breads = 15.0
    _sell(client, token, bread_id, qty=3, alice_id=alice_id, pay_later=True)

    resp = client.get("/reports", headers=_auth(token))
    assert resp.status_code == 200, resp.text
    data = resp.json()

    ov = data["overview"]
    assert ov["sales_count"] == 2
    assert ov["paid_revenue"] == 10.0
    assert ov["unpaid_debt"] == 15.0
    assert ov["grand_total"] == 25.0


def test_report_sales_details_present(client):
    """sales_details has one entry per SaleItem."""
    token = _login(client)
    _, bread_id, alice_id = _setup_world(client, token)
    _sell(client, token, bread_id, qty=2)

    resp = client.get("/reports", headers=_auth(token))
    assert resp.status_code == 200
    details = resp.json()["sales_details"]
    assert len(details) == 1
    assert details[0]["product"] == "bread"
    assert details[0]["qty"] == 2
    assert details[0]["line_total"] == 10.0
    assert details[0]["payment_status"] == "Paid"


def test_report_inventory_consumption_reflects_recipe(client):
    """inventory_consumption aggregates recipe×qty for sold products."""
    token = _login(client)
    flour_id, bread_id, _ = _setup_world(client, token)

    # Sell 3 bread; recipe: bread <- flour qty=2 → consumed = 6
    _sell(client, token, bread_id, qty=3)

    resp = client.get("/reports", headers=_auth(token))
    assert resp.status_code == 200
    consumption = resp.json()["inventory_consumption"]
    assert len(consumption) == 1
    entry = consumption[0]
    assert entry["ingredient"] == "flour"
    assert entry["consumed"] == 6.0  # 2 * 3
    # Remaining: 200 - 6 = 194
    assert entry["remaining"] == 194.0
    assert entry["unit"] == "kg"


def test_report_current_inventory_present(client):
    """current_inventory lists all inventory items ordered by name."""
    token = _login(client)
    flour_id, _, _ = _setup_world(client, token)

    resp = client.get("/reports", headers=_auth(token))
    assert resp.status_code == 200
    inv = resp.json()["current_inventory"]
    names = [i["name"] for i in inv]
    assert "flour" in names


def test_report_customer_summary_aggregates_paid_vs_debt(client):
    """customer_summary correctly aggregates a customer's paid vs debt."""
    token = _login(client)
    _, bread_id, alice_id = _setup_world(client, token)

    # Paid: 1 bread = 5.0
    _sell(client, token, bread_id, qty=1, alice_id=alice_id)
    # Unpaid: 2 breads = 10.0
    _sell(client, token, bread_id, qty=2, alice_id=alice_id, pay_later=True)

    resp = client.get("/reports", headers=_auth(token))
    assert resp.status_code == 200
    summary = resp.json()["customer_summary"]
    assert len(summary) == 1
    alice = summary[0]
    assert alice["customer"] == "Alice"
    assert alice["purchases"] == 2
    assert alice["paid"] == 5.0
    assert alice["debt"] == 10.0


def test_report_unpaid_bills_lists_unpaid_sale(client):
    """unpaid_bills includes the pay_later sale."""
    token = _login(client)
    _, bread_id, alice_id = _setup_world(client, token)

    paid_sale = _sell(client, token, bread_id, qty=1, alice_id=alice_id)
    unpaid_sale = _sell(client, token, bread_id, qty=4, alice_id=alice_id, pay_later=True)

    resp = client.get("/reports", headers=_auth(token))
    assert resp.status_code == 200
    unpaid = resp.json()["unpaid_bills"]
    unpaid_ids = [u["sale_id"] for u in unpaid]
    assert unpaid_sale["sale_id"] in unpaid_ids
    assert paid_sale["sale_id"] not in unpaid_ids


# ---------------------------------------------------------------------------
# Date range filtering
# ---------------------------------------------------------------------------

def test_report_custom_range_future_returns_empty(client):
    """A custom range in the future yields zero sales."""
    token = _login(client)
    _, bread_id, _ = _setup_world(client, token)
    _sell(client, token, bread_id, qty=1)

    resp = client.get(
        "/reports",
        params={"start": "2099-01-01", "end": "2099-12-31"},
        headers=_auth(token),
    )
    assert resp.status_code == 200
    ov = resp.json()["overview"]
    assert ov["sales_count"] == 0
    assert ov["paid_revenue"] == 0.0
    assert ov["grand_total"] == 0.0


def test_report_bad_date_format_returns_400(client):
    """A malformed date string in start/end raises HTTP 400."""
    token = _login(client)

    resp = client.get(
        "/reports",
        params={"start": "not-a-date", "end": "2099-12-31"},
        headers=_auth(token),
    )
    assert resp.status_code == 400


def test_report_bad_end_date_returns_400(client):
    """A malformed end date string raises HTTP 400."""
    token = _login(client)

    resp = client.get(
        "/reports",
        params={"start": "2099-01-01", "end": "31/12/2099"},
        headers=_auth(token),
    )
    assert resp.status_code == 400


# ---------------------------------------------------------------------------
# Role access
# ---------------------------------------------------------------------------

def test_salesman_can_access_reports(client):
    """salesman role is allowed to access /reports."""
    admin_token = _login(client)
    _, bread_id, _ = _setup_world(client, admin_token)
    _create_user(client, admin_token, "sally", "salesman")
    sally_token = _login(client, "sally", "pass")

    resp = client.get("/reports", headers=_auth(sally_token))
    assert resp.status_code == 200


def test_stockman_cannot_access_reports(client):
    """stockman role is forbidden from /reports."""
    admin_token = _login(client)
    _create_user(client, admin_token, "stocky", "stockman")
    stocky_token = _login(client, "stocky", "pass")

    resp = client.get("/reports", headers=_auth(stocky_token))
    assert resp.status_code == 403


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

EXPECTED_SHEETS = [
    "Summary",
    "Sales Details",
    "Inventory Consumption",
    "Current Inventory",
    "Customer Summary",
    "Unpaid Bills",
]


def test_export_xlsx_returns_200_with_correct_content_type(client):
    """GET /reports/export.xlsx returns 200 and xlsx content-type."""
    token = _login(client)
    _, bread_id, alice_id = _setup_world(client, token)
    _sell(client, token, bread_id, qty=2)
    _sell(client, token, bread_id, qty=1, alice_id=alice_id, pay_later=True)

    resp = client.get("/reports/export.xlsx", headers=_auth(token))
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]
    assert len(resp.content) > 0


def test_export_xlsx_has_expected_sheets(client):
    """The exported workbook contains all required sheet names."""
    token = _login(client)
    _, bread_id, alice_id = _setup_world(client, token)
    _sell(client, token, bread_id, qty=2)
    _sell(client, token, bread_id, qty=3, alice_id=alice_id, pay_later=True)

    resp = client.get("/reports/export.xlsx", headers=_auth(token))
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    for sheet_name in EXPECTED_SHEETS:
        assert sheet_name in wb.sheetnames, f"Sheet '{sheet_name}' missing from workbook"


def test_export_xlsx_summary_sheet_has_data(client):
    """The Summary sheet in the exported workbook contains the overview rows."""
    token = _login(client)
    _, bread_id, _ = _setup_world(client, token)
    _sell(client, token, bread_id, qty=2)

    resp = client.get("/reports/export.xlsx", headers=_auth(token))
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb["Summary"]

    # Row 1 = header, rows 2-5 = data
    metrics = [ws.cell(row=i, column=1).value for i in range(2, 6)]
    assert "Sales Count" in metrics
    assert "Paid Revenue" in metrics
    assert "Grand Total" in metrics


def test_export_xlsx_stockman_forbidden(client):
    """stockman cannot access the xlsx export endpoint either."""
    admin_token = _login(client)
    _create_user(client, admin_token, "stocky2", "stockman")
    stocky_token = _login(client, "stocky2", "pass")

    resp = client.get("/reports/export.xlsx", headers=_auth(stocky_token))
    assert resp.status_code == 403


def test_export_xlsx_empty_report_still_has_sheets(client):
    """Even with no sales the xlsx still has all required sheets."""
    token = _login(client)

    resp = client.get("/reports/export.xlsx", headers=_auth(token))
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    for sheet_name in EXPECTED_SHEETS:
        assert sheet_name in wb.sheetnames
